import streamlit as st
import openpyxl
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import anthropic
import io
import json
import re
from typing import Optional

st.set_page_config(
    page_title="Excel 번역기",
    page_icon="📊",
    layout="wide",
)

st.title("📊 Excel 번역기")
st.caption("Excel 파일의 텍스트를 영어로 번역합니다. 원본 서식(폰트, 색상, 테두리, 병합, 너비 등)이 유지됩니다.")

# ─── 번역 함수 ────────────────────────────────────────────────────────────────

MODELS = {
    "claude-haiku-4-5 (빠름/저렴)": "claude-haiku-4-5-20251001",
    "claude-sonnet-4-6 (균형)": "claude-sonnet-4-6",
    "claude-opus-4-6 (고품질)": "claude-opus-4-6",
}


def batch_translate(client: anthropic.Anthropic, texts: list[str], model: str) -> list[str]:
    """Claude API로 텍스트 배치 번역 (JSON 입출력)"""
    if not texts:
        return []

    texts_json = json.dumps(texts, ensure_ascii=False)

    prompt = (
        "Translate the following texts to English.\n"
        "Return ONLY a valid JSON array with the exact same number of elements.\n"
        "Preserve newlines (\\n) and special characters inside each string.\n"
        "If a text is already in English, keep it unchanged.\n"
        "Do NOT add explanations or markdown formatting.\n\n"
        f"Input:\n{texts_json}\n\nOutput:"
    )

    try:
        message = client.messages.create(
            model=model,
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}],
        )
        response_text = message.content[0].text.strip()

        # 마크다운 코드블록 제거
        response_text = re.sub(r"^```(?:json)?\n?", "", response_text)
        response_text = re.sub(r"\n?```$", "", response_text)

        translated = json.loads(response_text)

        if not isinstance(translated, list) or len(translated) != len(texts):
            st.warning(
                f"번역 결과 개수 불일치 (입력 {len(texts)}개 → 출력 {len(translated) if isinstance(translated, list) else '?'}개). "
                "해당 배치는 원본 유지됩니다."
            )
            return texts

        return [str(t) for t in translated]

    except json.JSONDecodeError as e:
        st.error(f"JSON 파싱 오류: {e}\n응답: {response_text[:300]}")
        return texts
    except anthropic.APIError as e:
        st.error(f"API 오류: {e}")
        return texts


def translate_worksheet(
    ws,
    client: anthropic.Anthropic,
    model: str,
    cell_range: Optional[str],
    progress_bar,
    batch_size: int = 30,
) -> int:
    """워크시트의 텍스트 셀을 번역하고 번역된 셀 수 반환"""

    # 순회할 행 결정
    if cell_range:
        try:
            rows = ws[cell_range]
            # 단일 셀 또는 단일 행이면 튜플로 감싸기
            if not isinstance(rows, (list, tuple)):
                rows = ((rows,),)
            elif rows and not isinstance(rows[0], (list, tuple)):
                rows = (rows,)
        except Exception as e:
            st.error(f"범위 오류: {e}")
            return 0
    else:
        rows = list(ws.iter_rows())

    # 번역 대상 셀 수집
    text_cells = []
    for row in rows:
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if not isinstance(cell.value, str):
                continue
            if not cell.value.strip():
                continue
            if cell.value.startswith("="):
                continue
            text_cells.append(cell)

    total = len(text_cells)
    if total == 0:
        progress_bar.progress(1.0, text="번역할 텍스트 없음")
        return 0

    translated_count = 0

    for i in range(0, total, batch_size):
        batch = text_cells[i : i + batch_size]
        texts = [cell.value for cell in batch]

        translated = batch_translate(client, texts, model)

        for cell, new_text in zip(batch, translated):
            cell.value = new_text

        translated_count += len(batch)
        pct = translated_count / total
        progress_bar.progress(pct, text=f"{translated_count} / {total} 셀 번역 완료")

    return total


# ─── 사이드바 설정 ────────────────────────────────────────────────────────────

with st.sidebar:
    st.header("⚙️ 설정")

    api_key = st.text_input(
        "Anthropic API Key",
        type="password",
        help="https://console.anthropic.com 에서 발급",
    )

    model_label = st.selectbox("모델 선택", list(MODELS.keys()), index=1)
    selected_model = MODELS[model_label]

    batch_size = st.slider(
        "배치 크기",
        min_value=5,
        max_value=80,
        value=30,
        help="한 번에 번역할 셀 수. 클수록 API 호출 횟수가 줄지만 오류 위험 증가",
    )

    st.divider()
    st.markdown(
        "**사용 방법**\n"
        "1. API Key 입력\n"
        "2. Excel 파일 업로드\n"
        "3. 번역할 시트/범위 선택\n"
        "4. '번역 시작' 클릭\n"
        "5. 완료 후 파일 다운로드"
    )

# ─── 메인 영역 ────────────────────────────────────────────────────────────────

uploaded_file = st.file_uploader(
    "Excel 파일 업로드 (.xlsx)",
    type=["xlsx"],
    help=".xlsx 형식만 지원합니다. .xls는 먼저 .xlsx로 변환해 주세요.",
)

if uploaded_file:
    try:
        wb_preview = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet_names = wb_preview.sheetnames
        wb_preview.close()

        col1, col2 = st.columns(2)

        with col1:
            selected_sheets = st.multiselect(
                "번역할 시트",
                options=sheet_names,
                default=sheet_names,
            )

        with col2:
            use_range = st.checkbox("특정 범위만 번역 (선택사항)")
            cell_range = None
            if use_range:
                cell_range = st.text_input(
                    "셀 범위",
                    placeholder="예: A1:F30",
                    help="모든 선택된 시트에 동일한 범위가 적용됩니다.",
                )
                if cell_range and not re.match(
                    r"^[A-Za-z]+\d+:[A-Za-z]+\d+$", cell_range
                ):
                    st.error("올바른 범위 형식: A1:D10")
                    cell_range = None

        st.divider()

        if not selected_sheets:
            st.warning("번역할 시트를 하나 이상 선택하세요.")
        elif not api_key:
            st.warning("사이드바에서 Anthropic API Key를 입력하세요.")
        else:
            if st.button("🚀 번역 시작", type="primary", use_container_width=True):
                uploaded_file.seek(0)
                wb_work = load_workbook(uploaded_file, data_only=False)

                client = anthropic.Anthropic(api_key=api_key)
                total_cells = 0

                for sheet_name in selected_sheets:
                    st.write(f"**📄 시트: {sheet_name}**")
                    prog = st.progress(0, text="준비 중...")

                    ws = wb_work[sheet_name]
                    count = translate_worksheet(
                        ws, client, selected_model, cell_range, prog, batch_size
                    )
                    total_cells += count
                    prog.progress(1.0, text=f"완료 — {count}개 셀 번역됨")

                st.success(f"✅ 번역 완료! 총 **{total_cells}**개 셀 번역")

                # 결과 저장
                output = io.BytesIO()
                wb_work.save(output)
                output.seek(0)

                base_name = uploaded_file.name.rsplit(".", 1)[0]
                st.download_button(
                    label="📥 번역된 파일 다운로드",
                    data=output,
                    file_name=f"{base_name}_translated.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

    except Exception as e:
        st.error(f"파일 로드 오류: {e}")
